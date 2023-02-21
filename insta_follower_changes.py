import os
import sys
import openpyxl
from instagrapi import Client
from pwinput import pwinput
from tempfile import gettempdir
from datetime import datetime
from pathlib import Path
import requests
import discord
from dotenv import load_dotenv
load_dotenv()
client = discord.Client(intents=discord.Intents.default())
@client.event
async def on_ready():
    print('Bot is ready.')
client.run(os.getenv('DISCORD_BOT_TOKEN'))


# Get user information
class following_main:
    answer= input("\nWould you like to use existing account? (y/n): ").strip().lower()
    if answer == "y":
        print("Displaying existing accounts: ")
        [print(i) for i in os.listdir(gettempdir()) if i.endswith("_account_dump.json")]
        print("\n-----","\nPlease enter the username of the account you would like to use: ")
        login_dict = {
                        os.getenv("TRH1U"): os.getenv("TRH1"),
                        os.getenv("BRITTJAMSU"): os.getenv("BRITTJAMS"),
                        os.getenv("BOBU"): os.getenv("BOB"),
                    }                                              
        username = input("Username: ").strip().lower()
        if username in login_dict.keys():
            password = login_dict.get(username)
            oauth = input("2FA Code (if not exists, leave blank): ").strip()
    elif answer == "n":
        print("-----")
        username = input("Username: ").strip()
        password = pwinput()
        oauth = input("2FA Code (if not exists, leave blank): ").strip()
        print()
    

    # Account settings dump path
    cwd = os.getcwd()
    DUMP = os.path.join(gettempdir(), "{}_account_dump.json".format(username))
    BACKUP = os.path.join(str(Path.home()), "{}_follower_dump.instabackup".format(username))
    # Try to login and get account details, if fails, terminate
    try:
                # Create the client
        cl = Client()

        if username == os.getenv("TRH1U"):
            cl.set_proxy(os.getenv("PROXYZ1"))
        elif username == os.getenv("BRITTJAMSU"):
            cl.set_proxy(os.getenv("PROXYINS"))
        elif username == (os.getenv("BOBU")):
            cl.set_proxy(os.getenv("PROXYDIND"))
        else:
            cl.set_proxy(os.getenv("PROXYZ1"))
        # If there is a dump for the account, load settings
        if os.path.exists(DUMP):
            cl.load_settings(DUMP)
            print("Loaded account settings from dump.")

        # Login
        cl.login(username, password, verification_code=oauth)
        # If there is not a dump for the account, dump settings
        if not os.path.exists(DUMP):
            cl.dump_settings(DUMP)
            print("Dumped account settings to temp directory.")
        # Get the user ID
        user_id = cl.user_id
        print("Your User ID: {}".format(user_id))
        

        # Get following
        print("Getting accounts that you follow...")
        following = cl.user_following(user_id)
        print("You are following {} accounts.".format(len(following)))
                
    except Exception as e:
        # Print the exception message
        print("An exception occured: {}".format(str(e)))

        # Delete account dump to prevent login issues
        if os.path.exists(DUMP):
            os.remove(DUMP)
            print("Account dump is deleted to prevent login issues.")
        # Terminate
        input("Press Enter to terminate.")
        sys.exit(1)
                
            # Get the path for results
           
    path0= os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop', "{}_follower_changes.html".format(username))
    path1 = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop', "{}_following_all.html".format(username))
    path = os.path.join(cwd, "{}_following_all_table.html".format(username))
            
            
            

            # Print the new following and fill an HTML body with them and their Instagram profiles

    html = "<b>Timestamp:</b> {}<br><br>".format(str(datetime.now()))
    html2 = "<b>Timestamp:</b> {}<br><br>".format(str(datetime.now()))
    html2 += "<table border=1>"

    # Extract only the usernames on both followed accounts and followers
    following_dict = {}

    for i in following.values():
        following_dict[i.pk] = i.username

    # If there is no backup, create the backup and terminate
    if not os.path.exists(BACKUP):
        print("Backup not found. Creating a new backup...")
        print("Saving Following list to {}".format(path1))
        html += "<b>following:</b><br>"
        html2 += "<tr> <th>following:</th> </tr><br>"
        for i in following_dict.values():
            html += '<a href="https://instagram.com/{}">{}</a><br><br>'.format(i, i)
            html2 += '<tr><td> <a href="https://instagram.com/{}">{}</a></td></tr>'.format(i, i)
        html2 += "</table>"    
        with open(path1, "w") as file:
            file.write(html)
        with open(path, "w") as file:
            file.write(html2)
        # Create the backup file
        with open(BACKUP, "w") as file:
            file.write(str(datetime.now()) + "\n" + "\n".join([";".join(i) for i in following_dict.items()]))
        print("Backup created.")
        # Terminate
        input("Press Enter to terminate.")
        sys.exit(0)

    # Get backup values
    print("Getting backup...")
    with open(BACKUP, "r") as file:
        lines = file.readlines()
    timestamp = lines.pop(0).strip()
    old_backup = {}
    for i in lines:
        item = i.strip().split(";")
        old_backup[item[0]] = item[1]

    print("Latest backup timestamp: " + timestamp) # Print the backup timestamp
    
    def following_status(following_dict, old_backup, path1, path, html, html2,path0,BACKUP):
        # Following change lists
        new_following = []
        unfollowed = []

        # Get new following
        for i in following_dict.items():
            if i[0] not in old_backup.keys():
                new_following.append(i[1])

        # Get users who have unfollowed
        for i in old_backup.items():
            if i[0] not in following_dict.keys():
                unfollowed.append(i[1])

        print()
        # If there are no changes, terminate
        if len(new_following) == 0 and len(unfollowed) == 0:
            print("No new changes on followers.")
            print("following List saved to {}".format(path1))
            html += "<b>following:</b><br>"
            html2 += "<tr> <th>following:</th></tr>"
            for i in following_dict.values():
                html += '<a href="https://instagram.com/{}">{}</a><br>'.format(i, i)
                html2 += '<tr><td><a href="https://instagram.com/{}">{}</a></td></tr>'.format(i, i)
            with open(path1, "w") as file:
                file.write(html)
            with open(path, "w") as file:
                file.write(html2)
            input("Press Enter to terminate.")
            sys.exit(0)
                
                
        if len(new_following) != 0:
            print("New follows:")
            print("\n".join(new_following))
            html += "<b>New follows:</b><br>"
            html2 += "<tr><th>New follows:</th></tr></br>"
            for i in new_following:
                html += '<a href="https://instagram.com/{}">{}</a><br>'.format(i, i)
                html2 += '<td><a href="https://instagram.com/{}">{}</a></td>'.format(i, i)
            if len(unfollowed) != 0:
                print()
                html += "<br>"

        # Print the unfollowed users and fill an HTML body with them and their Instagram profiles
        if len(unfollowed) != 0:
            print("Unfollowed:")
            print("\n".join(unfollowed))
            html += "<b>Unfollowed:</b><br>"
            html2 += "<tr><th>Unfollowed:</th></tr></br>"
            for i in unfollowed:
                html += '<a href="https://instagram.com/{}">{}</a><br>'.format(i, i)

        # Open the file and write the results into it
        #create a new file to write the new following list
        print("Saving Following list to {}".format(path1))
        html += "<br><b>following:</b><br>"
        html2 += "<tr><th>following:</th></tr>"
        for i in following_dict.values():
            html += '<a href="https://instagram.com/{}">{}</a><br>'.format(i, i)
            html2 += '<td><a href="https://instagram.com/{}">{}</a></td>'.format(i, i)
            html2 +="</table>"
        with open(path0, "a") as file: 
            file.write(html)
        with open(path1, "a") as file:
            file.write(html)
        with open(path, "a") as file:
            file.write(html2)

        # Backup again
        with open(BACKUP, "w") as file:
            file.write(str(datetime.now()) + "\n" + "\n".join([";".join(i) for i in following_dict.items()]))

        # Print the path and terminate
        print("\nResults are saved to {}.".format(path0, path1, path))
        input("Press Enter to terminate.")
    def options():
         yes_choices = ['yes', 'y']
         no_choices = ['no', 'n']
         api_options = ['media', 'names','following-status']
         while True:
            user_input = input('Do you want to continue?[media,following-status,names,(no/n)]: ')
            if user_input.lower() in api_options:
                print('okay...{}'.format(user_input))
                return user_input
            elif user_input.lower() in no_choices:
                print('goodbye')
                SystemExit(0)
                break
            else:
                print('Type yes or no')
                continue
    anw=options()
    if anw=='following-status':
        following_status(following_dict, old_backup, path1, path, html, html2,path0,BACKUP)
    elif anw=='names':
        following = cl.user_following(cl.user_id)
         # Extract full names from user objects
        full_names = [user.full_name for user in following.values()]

        # Create a new workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Write header and full names to worksheet
        worksheet.cell(row=1, column=1, value="Full Name")
        for i, name in enumerate(full_names):
            worksheet.cell(row=i+2, column=1, value=name)
        # Save the workbook
        workbook.save("full_names_{}.xlsx".format(username))
        print(full_names)
    elif anw=='media':
            following=cl.user_following(cl.user_id)
            media = []
            for user_id in following:
                user_media = cl.user_medias(cl.user_id,amount=1)
            for m in user_media:
                media.append(m)
            webhook_url = os.getenv("DISCORD_WEBHOOK_URL")
            for m in media:
                data = {"content": m.url}
                headers = {"Content-Type": "application/json"}
                r = requests.post(webhook_url, json=data, headers=headers)
            if r.status_code != 204:
                print(f"An error occurred while sending the message to the Discord channel: {r.text}")
            else:
                print(f"Media {m.pk} sent to the Discord channel.")
    print("All media sent to the Discord channel.")
    input("Press Enter to terminate.")
    # Send the media file to the Discord channel
    

       


     
       
    